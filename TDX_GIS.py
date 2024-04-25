"""
!/usr/bin/env python
-*- coding: utf-8 -*-
@File  : TDX_GIS.py
@Author: GinaChou
@Date  : 2023/11/30
"""
import requests
from pprint import pprint
import json
import os
import openpyxl

app_id = 'ginachou-1524620d-b8ae-4796'
app_key = 'e0c769a4-0796-4540-b504-d66668428ce1'

auth_url="https://tdx.transportdata.tw/auth/realms/TDXConnect/protocol/openid-connect/token"
url = "https://tdx.transportdata.tw/api/basic/v2/Road/Link/Shape/Geometry/WKT/0000300140000T?%24format=JSON"


class Auth():

    def __init__(self, app_id, app_key):
        self.app_id = app_id
        self.app_key = app_key

    def get_auth_header(self):
        content_type = 'application/x-www-form-urlencoded'
        grant_type = 'client_credentials'

        return{
            'content-type' : content_type,
            'grant_type' : grant_type,
            'client_id' : self.app_id,
            'client_secret' : self.app_key
        }

class data():

    def __init__(self, app_id, app_key, auth_response):
        self.app_id = app_id
        self.app_key = app_key
        self.auth_response = auth_response

    def get_data_header(self):
        auth_JSON = json.loads(self.auth_response.text)
        access_token = auth_JSON.get('access_token')

        return{
            'authorization': 'Bearer '+access_token
        }

if __name__ == '__main__':
    url2 = 'https://tdx.transportdata.tw/api/basic/v2/Road/Link/Shape/Geometry/WKT/'
    params = {
        '$format': 'JSON',
    }
    # json_data = [
    #     '0000300140000T',
    #     '0000301168000T',
    # ]
    json_data = [
        '0000210000000H',
        '0000210000062H',
    ]

    try:
        d = data(app_id, app_key, auth_response)
        data_response = requests.get(url, headers=d.get_data_header())
    except:
        a = Auth(app_id, app_key)
        auth_response = requests.post(auth_url, a.get_auth_header())
        d = data(app_id, app_key, auth_response)
        data_response = requests.get(url, headers=d.get_data_header())
        res_json = requests.post(url2, params=params,headers=d.get_data_header(),json=json_data)

    print(f'auth:{auth_response}') #AUTH標頭
    pprint(f'auth_text:{auth_response.text}')
    # print(f'res_json:{res_json}')
    # pprint(f'res_json_text:{res_json.text}')  ##POST產出

    res_list = json.loads(res_json.text)
    # pprint(f'data_list: {res_list}')
    # print(f'data_type: {type(res_list)}')

    ##切換到.py檔目前的絕對路徑，並於底下新增output資料夾
    result_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),'output')
    ##檢查是否存在output資料夾，若無則新增
    if not os.path.exists(result_path):
        os.makedirs(result_path)
    # # os.chdir 是 python 切換到電腦指定路徑的方法
    os.chdir(result_path)

    # wb = openpyxl.load_workbook('oxxo.xlsx', data_only=True)
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet("POST", 0) # 新增工作表 1
    result = []
    ##為EXCEL增加標頭
    titles = ("LinkID","Geometry","Version","UpdateDate")
    # titles = ("LinkID","Version")
    sheet.append(titles)
    # 輸出link資訊(分列)
    for link in res_list:
        ls = []
        LinkID = link["LinkID"]
        Geo = link["Geometry"]
        # print("Geometry:", link["Geometry"])
        ##只保留經緯度資料
        Geometry = Geo.replace("LINESTRING","")
        Geometry = Geometry.replace("(","")
        Geometry = Geometry.replace(")","")
        Geometry = Geometry.replace(",",";") ##逗號改為分號
        # print("Geometry(修):", Geometry)
        Version = link["Version"]
        UpdateDate = link["UpdateDate"]
        ls.append(LinkID)
        ls.append(Geometry)
        ls.append(Version)
        ls.append(UpdateDate)

        result.append(ls)
        print("result:", result)

        sheet.append(ls)

    print("FINAL_result:", result)
    wb.save("Geometry.xlsx")

#     ############GET(OK)#########################
#     print(f'text_type: {type(data_response.text)}')
#     data_text = data_response.text
#     data_list = json.loads(data_text)
#     pprint(f'data_list: {data_list}')
#     print(f'data_type: {type(data_list)}')
#     # 輸出link資訊
#     for link in data_list:
#         print("LinkID:", link["LinkID"])
#         print("Geometry:", link["Geometry"])
#         print("Version:", link["Version"])
#         print("UpdateDate:", link["UpdateDate"])
# ######################################

# 这是一个示例 Python 脚本。
# 按 Shift+F10 执行或将其替换为您的代码。
# 按 双击 Shift 在所有地方搜索类、文件、工具窗口、操作和设置。
import os
import time
import math
import requests
import json
# from openpyxl import load_workbook
import openpyxl
import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app_id = 'ginachou-1524620d-b8ae-4796'
app_key = 'e0c769a4-0796-4540-b504-d66668428ce1'
auth_url = "https://tdx.transportdata.tw/auth/realms/TDXConnect/protocol/openid-connect/token"
url_geo = 'https://tdx.transportdata.tw/api/basic/v2/Road/Link/Shape/Geometry/WKT/'
# 'https://tdx.transportdata.tw/api/basic/v2/Road/Link/LinkID/'


def print_hi(name):
    # 在下面的代码行中使用断点来调试脚本。
    print(f'Hi, {name}')  # 按 Ctrl+F8 切换断点。


class Auth():

    def __init__(self, app_id, app_key):
        self.app_id = app_id
        self.app_key = app_key

    def get_auth_header(self):
        content_type = 'application/x-www-form-urlencoded'
        grant_type = 'client_credentials'

        return{
            'content-type': content_type,
            'grant_type': grant_type,
            'client_id': self.app_id,
            'client_secret': self.app_key
        }


class data():

    def __init__(self, app_id, app_key, auth_response):
        self.app_id = app_id
        self.app_key = app_key
        self.auth_response = auth_response

    def get_data_header(self):
        auth_JSON = json.loads(self.auth_response.text)
        access_token = auth_JSON.get('access_token')

        return{
            'authorization': 'Bearer ' + access_token
        }


# 按间距中的绿色按钮以运行脚本。
if __name__ == '__main__':
    params = {
        '$format': 'JSON',
    }
    dir = '路段編碼.xlsx'
    wb = openpyxl.load_workbook(dir)
    # 獲取所有表名，得到工作簿的所有工作表名 結果： ['Sheet1', 'Sheet2', 'Sheet3']
    sheet_names = wb.sheetnames
    # sheet1 = wb[sheet_names[0]]  # 打開第一個 sheet 工作表
    # 根據表名打開sheet表
    sheet = wb['缺']  # 取得工作表名稱為「 」的內容

    # 獲取C列的所有資料
    list_column_C = []  # 請求參數
    for e in sheet["C"]:
        list_column_C.append(e.value)
        # print(f'excel_e: {e.value}')
    row = 2
    i = 2
    # print(f'excel_ID: {e.value}\nexcel_type: {type(e.value)}')
    for str_param in list_column_C[1:]:
        # 還要看到底是手動增加LINKID還是用TDX  str_param.strip().split(",")
        json_data = str_param.strip().split(",")
        print(f'json_ID: {json_data}\njson_type: {type(json_data)}')
        try:
            a = Auth(app_id, app_key)
            auth_response = requests.post(auth_url, a.get_auth_header())
            d = data(app_id, app_key, auth_response)
            res_geo = requests.post(
                url_geo,
                params=params,
                headers=d.get_data_header(),
                json=json_data)
            # print(f'res_geo: {res_geo}\nres_geo: {type(res_geo)}')
            res_list = json.loads(res_geo.text)  # 這邊順序就變了 2023/12/18
            print(f'res_geo: {res_list}')
        except BaseException:
            logger.error('《' + str(row) + '》項，介面地址或入參異常！！！')
        row += 1

        ##為EXCEL增加標頭  應該用不到
        titles = ("LinkID","Geometry","Version","UpdateDate")
        # sheet.append(titles)  ##要改成在已存在檔案中寫入資料
        LinkID = ""
        Version = ""
        UpdateDate = ""
        Geometry_result = []
###20240425######
##已處理首尾重複,用移除重複值容易發生四捨五入一樣被誤刪的事故
##已處理浮點數五位數末碼為0會消失議題
        index = 1  ##路段的第幾組LINKID
        for link in res_list:
            Geo = link["Geometry"]
            tempGeo = Geo.replace("LINESTRING(", "")
            tempGeo_s2 = tempGeo.replace(")", "")
            tempGeo_s3= tempGeo_s2.split(",")
            print(f'index: {index}')
            print(f'link: {link}')
            print(f'->tempGeo_s3: {tempGeo_s3}')
            # print(f'\ntype: {type(tempGeo_s3)}')
            ## 迭代link清單中的元素
            for k, point in enumerate(tempGeo_s3):
                print(f'k:{k}')
                if index != 1 and k == 0:##路段非第一組LINKID路線的第一個經緯度
                    print(f'路段中後續的LINID不加入第一個經緯度')
                    continue
                else:
                    tempGeo_s4 = point.split(" ")
                    tempLon1 = float(tempGeo_s4[0])
                    tempLon1 = format(round(tempLon1, 5),'.5f') ##保持輸出格式為浮點數五位數，末碼為0不會消失
                    tempLat1 = float(tempGeo_s4[1])
                    tempLat1 = format(round(tempLat1, 5),'.5f')
                    point = str(tempLon1) + " " + str(tempLat1)
                    # print(f'point:{point}')
                    Geometry_result.append(point)
                    print(f'-->Geometry_result:{Geometry_result}')
                    # print(f'\ntype: {type(Geometry_result)}')
            geomtry =','.join(Geometry_result) #依迴圈Geometry_result分別加入WKT，以,分隔
            geomtry = "LINESTRING("+geomtry+")" ##加入頭尾
            print(f'==>Geometry:{geomtry}')
            Version = link["Version"]
            UpdateDate = link["UpdateDate"]
            index += 1 #路段群中的第幾個linkid
        ####s1.cell(1,2).value = 100     # 儲存格 B1 內容 ( row=1, column=2 ) 為 100
        # 寫入位置的行列號可以任意改變，這裡從第2行開始按行依次插入第4列
        sheet.cell(row=i, column=4).value = geomtry
        print(f'Geometry:{geomtry}')
        i += 1

    #     wb.save("路段編碼.xlsx")
###################################################################
    try:
        wb.save(dir)
        logger.info('測試資料保存成功！！！')
    except BaseException:
        logger.error('保存失敗，可能Excel檔案未關閉，請關閉Excel檔案後重新測試')

# 访问 https://www.jetbrains.com/help/pycharm/ 获取 PyCharm 帮助
# https://blog.csdn.net/l734971107/article/details/109635668
#https://blog.kyomind.tw/beyond-the-basic-stuff-with-python-02/
#https://blog.csdn.net/wg2627/article/details/126882315
#https://marco79423.net/articles/%E6%B7%BA%E8%AB%87-python-%E7%9A%84-for-%E8%BF%B4%E5%9C%88
#https://badgameshow.com/steven/python/%E4%BA%86%E8%A7%A3python%E4%B8%AD%E7%9A%84enumerate%E5%87%BD%E6%95%B8/

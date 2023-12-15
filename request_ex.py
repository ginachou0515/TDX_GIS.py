"""
!/usr/bin/env python
-*- coding: utf-8 -*-
@File  : request_ex.py
@Author: GinaChou
@Date  : 2023/12/11
"""
import os
import time
import requests
import json
from openpyxl  import load_workbook
import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ReadAPI():

    def ReadExel(self,dir):
        wb = load_workbook(dir)
        # 獲取所有表名
        sheet_names = wb.sheetnames  # 得到工作簿的所有工作表名 結果： ['Sheet1', 'Sheet2', 'Sheet3']
        # 根據表名打開sheet表
        sheet1 = wb[sheet_names[0]]  # 打開第一個 sheet 工作表

        # 獲取C列的所有資料
        list_sheet1_column_E = []  # 請求參數
        list_sheet1_column_B = []  # 請求地址
        list_sheet1_column_C = []  # 請求方式
        list_sheet1_column_A = []  # 請求名稱項
        list_sheet1_column_F = []  # 請求名稱項

        for b, e, c, a,f in zip(sheet1["B"], sheet1["E"], sheet1["C"], sheet1["A"],sheet1["F"]):
            list_sheet1_column_E.append(e.value)
            list_sheet1_column_B.append(b.value)
            list_sheet1_column_C.append(c.value)
            list_sheet1_column_A.append(a.value)
            list_sheet1_column_F.append(f.value)
        row = 2
        for url, method, str_param, testname,expect in zip(list_sheet1_column_B[1:], list_sheet1_column_C[1:], list_sheet1_column_E[1:], list_sheet1_column_A[1:],list_sheet1_column_F[1:]):
            try:
                ReadAPI.verdict(row,url, str_param,testname,sheet1,expect,method)  #讀寫excel
            except:
                logger.error('《' + str(testname) + '》項，介面地址或入參異常！！！')
                sheet1.cell(row=row, column=7, value='error')  # 響應結果
                sheet1.cell(row=row, column=8, value='error')  # 請求時間
                sheet1.cell(row=row, column=9, value='error')  # 狀態碼
                sheet1.cell(row=row, column=10, value='error')  # 判斷通過

            row+=1
        try:
            wb.save(dir)
            logger.info('測試資料保存成功！！！')
        except:
            logger.error('保存失敗，可能Excel檔案未關閉，請關閉Excel檔案後重新測試')

    def verdict(self,row,url, str_param,testname,sheet1,expect,method):

        if method=='get':
            param = eval(str_param)  # Exel讀出來的資料類型是字串，而get請求的入參必須是字典類型，post請求的入參是json字串類型
            API_data = ReadAPI.get(url, param, testname)
            ReadAPI.writedata(API_data,expect,sheet1,row,testname)

        elif method=='post':
            API_data = ReadAPI.post(url, str_param, testname)
            ReadAPI.writedata(API_data, expect, sheet1, row, testname)


    def get(self, url, param,testname):
        try:
            r = requests.get(url, params=param, timeout=1)
            r.raise_for_status()  # 如果響應狀態碼不是 200，就主動拋出異常
        except requests.RequestException as e:
            print(e)
        else:
            js = json.dumps(r.json())
            #print( '請求項名稱：'+testname+'、請求響應時間：'+str(r.elapsed.total_seconds()),'、請求狀態：'+str(r.status_code))
            return [r.json(), r.elapsed.total_seconds(),js,r.status_code]

    def post(self, url, param,testname):

        try:
            r = requests.post(url, data=param, timeout=1)
            r.raise_for_status()  # 如果響應狀態碼不是 200，就主動拋出異常
        except requests.RequestException as e:
            print(e)
        else:
            js = json.dumps(r.json())
            #print('請求項名稱：' + testname + '、請求響應時間：' + str(r.elapsed.total_seconds()), '、請求狀態：' + str(r.status_code))
            return [r.json(), r.elapsed.total_seconds(), js, r.status_code]

    def find_value(self,dir_data, fvalue):  #封裝個在查字典中的尋找value
        yesOrno = False
        for key in dir_data.keys():
            if (dir_data[key] == fvalue):

                yesOrno = True
                break

        return (yesOrno)

    def writedata(self,API_data,expect,sheet1,row,testname):
        API_data[2] = API_data[2].replace("'", '"')  # 替換"'", '"'
        dict_data = json.loads(API_data[2])  # 轉python字典
        if ReadAPI.find_value(dict_data, expect) == True:  # ReadAPI.get(url, param,testname)[3])==200:
            sheet1.cell(row=row, column=7, value=str(API_data[0]))  # 響應結果
            sheet1.cell(row=row, column=8, value=API_data[1])  # 請求時間
            sheet1.cell(row=row, column=9, value=int(API_data[3]))  # 狀態碼
            sheet1.cell(row=row, column=10, value="pass")  # 判斷通過
            logger.info('《' + str(testname) + '》項，響應通過、響應時間：' + str(API_data[1]) + '、狀態碼：' + str(API_data[3]))
        else:
            sheet1.cell(row=row, column=10, value="Fail")  # 判斷失敗
            sheet1.cell(row=row, column=8, value=API_data[1])  # 請求時間
            sheet1.cell(row=row, column=7, value=str(API_data[0]))  # 響應結果
            sheet1.cell(row=row, column=9, value=int(API_data[3]))  # 狀態碼
            logger.error('《' + str(testname) + '》項，響應錯誤、響應時間：' + str(API_data[1]) + '、狀態碼：' + str(API_data[3]))

if __name__ == '__main__':
    ReadAPI=ReadAPI()
    ##dir =os.path.join(os.getcwd()+'\介面測試用例.xlsx')
    dir ='介面測試用例.xlsx'

    ReadAPI.ReadExel(dir)


    input('Press Enter to exit...')


###https://blog.csdn.net/qq_42846555/article/details/97761484

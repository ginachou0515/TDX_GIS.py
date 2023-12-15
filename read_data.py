"""
!/usr/bin/env python
-*- coding: utf-8 -*-
@File  : read_data.py
@Author: GinaChou
@Date  : 2023/12/11
"""
##基本可以刪掉全部了，依EX修改##
import requests
import json
import os
import openpyxl
import time
from openpyxl  import load_workbook
import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app_id = 'ginachou-1524620d-b8ae-4796'
app_key = 'e0c769a4-0796-4540-b504-d66668428ce1'

auth_url="https://tdx.transportdata.tw/auth/realms/TDXConnect/protocol/openid-connect/token"
# url = "https://tdx.transportdata.tw/api/basic/v2/Road/Link/Shape/Geometry/WKT/0000300140000T?%24format=JSON"
url = 'https://tdx.transportdata.tw/api/basic/v2/Road/Link/LinkID/'
url2 = 'https://tdx.transportdata.tw/api/basic/v2/Road/Link/Shape/Geometry/WKT/'
url3 = 'https://tdx.transportdata.tw/api/basic/v2/Road/Link/LinkID/'

##請求方式
API_request= "post"
# # 請求名稱項
# API_request= []

class Auth():

    def __init__(self, app_id, app_key):
        self.app_id = app_id
        self.app_key = app_key

    def get_auth_header(self):
        content_type = 'application/x-www-form-urlencoded'
        grant_type = 'client_credentials'

        return{
            'content-type' : content_type,
            'grant_type' : grant_type,
            'client_id' : self.app_id,
            'client_secret' : self.app_key
        }

class data():

    def __init__(self, app_id, app_key, auth_response):
        self.app_id = app_id
        self.app_key = app_key
        self.auth_response = auth_response

    def get_data_header(self):
        auth_JSON = json.loads(self.auth_response.text)
        access_token = auth_JSON.get('access_token')

        return{
            'authorization': 'Bearer '+access_token
        }

class ReadAPI():

    def ReadExel(self,dir):
        wb = load_workbook(dir)
        # 獲取所有表名
        sheet_names = wb.sheetnames  # 得到工作簿的所有工作表名 結果： ['Sheet1', 'Sheet2', 'Sheet3']
        # 根據表名打開sheet表
        sheet1 = wb[sheet_names[0]]  # 打開第一個 sheet 工作表

        # 獲取C列的所有資料
        list_sheet1_column_N = []  # 請求參數
        # list_sheet1_column_B = []  # 請求地址
        # list_sheet1_column_C = []  # 請求方式
        # list_sheet1_column_A = []  # 請求名稱項
        # list_sheet1_column_F = []  # 請求名稱項

        for e in sheet1["N"]:
            list_sheet1_column_N.append(e.value)
            # list_sheet1_column_B.append(b.value)
            # list_sheet1_column_C.append(c.value)
            # list_sheet1_column_A.append(a.value)
            # list_sheet1_column_F.append(f.value)
        row = 2
        for method, str_param in zip(API_request, list_sheet1_column_N[1:]):
            try:
                ReadAPI.verdict(row,url, str_param,sheet1,method)  #讀寫excel
            except:
                logger.error('《' + row + '》項，介面地址或入參異常！！！')  ##8(L)Distance	9(M)Geometry 10(N)LinkID
                # sheet1.cell(row=row, column=7, value='error')  # 響應結果
                sheet1.cell(row=row, column=8, value='error')  # 請求時間
                sheet1.cell(row=row, column=9, value='error')  # 狀態碼
                # sheet1.cell(row=row, column=10, value='error')  # 判斷通過

            row+=1
        try:
            wb.save(dir)
            logger.info('測試資料保存成功！！！')
        except:
            logger.error('保存失敗，可能Excel檔案未關閉，請關閉Excel檔案後重新測試')

    def verdict(self,row,url, str_param,sheet1,method):

        if method=='get':
            param = eval(str_param)  # Exel讀出來的資料類型是字串，而get請求的入參必須是字典類型，post請求的入參是json字串類型
            # d = data(app_id, app_key, auth_response) ##2023/12/11
            API_data = ReadAPI.get(url, param)
            ReadAPI.writedata(API_data,sheet1,row)

        elif method=='post':
            # d = data(app_id, app_key, auth_response)  ##2023/12/11
            API_data = ReadAPI.post(url, str_param)
            ReadAPI.writedata(API_data, sheet1, row)


    def get(self, url, param):
        try:
            r = requests.get(url, params=param, timeout=1)
            r.raise_for_status()  # 如果響應狀態碼不是 200，就主動拋出異常
        except requests.RequestException as e:
            print(e)
        else:
            js = json.dumps(r.json())
            #print( '請求項名稱：'+testname+'、請求響應時間：'+str(r.elapsed.total_seconds()),'、請求狀態：'+str(r.status_code))
            return [r.json(), r.elapsed.total_seconds(),js,r.status_code]

    # def post(self, url, param,testname):
    #
    #     try:
    #         d = data(app_id, app_key, auth_response) ##2023/12/11
    #         # r = requests.post(url, data=param, timeout=1)
    #         r = requests.post(url,json=param, timeout=1, headers=d.get_data_header())
    #         r.raise_for_status()  # 如果響應狀態碼不是 200，就主動拋出異常
    #     except requests.RequestException as e:
    #         print(e)
    #     else:
    #         js = json.dumps(r.json())
    #         #print('請求項名稱：' + testname + '、請求響應時間：' + str(r.elapsed.total_seconds()), '、請求狀態：' + str(r.status_code))
    #         return [r.json(), r.elapsed.total_seconds(), js, r.status_code]

    def post(self, url, param):
        try:
            d = data(app_id, app_key, auth_response) ##2023/12/11
            r = requests.post(url,json=param, timeout=1, headers=d.get_data_header())
            r.raise_for_status()  # 如果響應狀態碼不是 200，就主動拋出異常
        except requests.RequestException as e:
            print(e)
        else:
            js = json.loads(r.json())
            print('、請求響應時間：' + str(r.elapsed.total_seconds()), '、請求狀態：' + str(r.status_code))
            print('json：' + r.json(), '、js：' + js)
            return [r.json(), r.elapsed.total_seconds(), js, r.status_code]


    def find_value(self,dir_data, fvalue):  #封裝個在查字典中的尋找value
        yesOrno = False
        for key in dir_data.keys():
            if (dir_data[key] == fvalue):

                yesOrno = True
                break

        return (yesOrno)

    def writedata(self,API_data,expect,sheet1,row,testname):
        API_data[2] = API_data[2].replace("'", '"')  # 替換"'", '"'
        dict_data = json.loads(API_data[2])  # 轉python字典
        if ReadAPI.find_value(dict_data, expect) == True:  # ReadAPI.get(url, param,testname)[3])==200:
            sheet1.cell(row=row, column=8, value=str(API_data[0]))  # 響應結果Distance
            # sheet1.cell(row=row, column=8, value=API_data[1])  # 響應結果
            # sheet1.cell(row=row, column=9, value=int(API_data[3]))  # 狀態碼
            # sheet1.cell(row=row, column=10, value="pass")  # 判斷通過
            logger.info('《' + str(row) + '》項，響應通過、響應時間：' + str(API_data[1]) + '、狀態碼：' + str(API_data[3]))
            # sheet1.cell(row=row, column=7, value=str(API_data[0]))  # 響應結果
            # sheet1.cell(row=row, column=8, value=API_data[1])  # 請求時間
            # sheet1.cell(row=row, column=9, value=int(API_data[3]))  # 狀態碼
            # sheet1.cell(row=row, column=10, value="pass")  # 判斷通過
            # logger.info('《' + str(testname) + '》項，響應通過、響應時間：' + str(API_data[1]) + '、狀態碼：' + str(API_data[3]))
        else:
            sheet1.cell(row=row, column=10, value="Fail")  # 判斷失敗
            sheet1.cell(row=row, column=8, value=API_data[1])  # 請求時間
            sheet1.cell(row=row, column=7, value=str(API_data[0]))  # 響應結果
            sheet1.cell(row=row, column=9, value=int(API_data[3]))  # 狀態碼
            logger.error('《' + str(testname) + '》項，響應錯誤、響應時間：' + str(API_data[1]) + '、狀態碼：' + str(API_data[3]))



if __name__ == '__main__':

    # params = {
    #     '$format': 'JSON',
    # }
    #
    # json_data = [
    #     '0000101006000F',
    #     '0000101007000F',
    #     '0000101007010F',
    # ]
    # print(f'json_data: {type(json_data)}')
    # try:
    #     d = data(app_id, app_key, auth_response)
    #     data_response = requests.get(url, headers=d.get_data_header())
    # except:
    #     a = Auth(app_id, app_key)
    #     auth_response = requests.post(auth_url, a.get_auth_header())
    #     d = data(app_id, app_key, auth_response)
    #     data_response = requests.get(url, headers=d.get_data_header())
    #     res_json = requests.post(url2, params=params,headers=d.get_data_header(),json=json_data)
    #     distance = requests.post(url3, params=params,headers=d.get_data_header(),json=json_data)
    # print(f'auth:{auth_response}') #AUTH標頭回應
    # pprint(f'auth_text:{auth_response.text}')
    #
    # # pprint(f'res_json_text:{res_json.text}')  ##POST產出
    # res_list = json.loads(res_json.text)
    # # print(f'data_type: {type(res_list)}')
    # dist_list = json.loads(distance.text)
    # print(f'geometry: {res_list}')
    # print(f'distance: {dist_list}')


    ReadAPI=ReadAPI()
    ##dir =os.path.join(os.getcwd()+'\介面測試用例.xlsx')
    dir ='介面測試.xlsx'

    ReadAPI.ReadExel(dir)

    input('Press Enter to exit...')
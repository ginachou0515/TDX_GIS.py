# 这是一个示例 Python 脚本。
# 按 Shift+F10 执行或将其替换为您的代码。
# 按 双击 Shift 在所有地方搜索类、文件、工具窗口、操作和设置。
import os
import time
import requests
import json
from openpyxl  import load_workbook
import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app_id = 'ginachou-1524620d-b8ae-4796'
app_key = 'e0c769a4-0796-4540-b504-d66668428ce1'
auth_url="https://tdx.transportdata.tw/auth/realms/TDXConnect/protocol/openid-connect/token"
url_len = 'https://tdx.transportdata.tw/api/basic/v2/Road/Link/LinkID/'

def print_hi(name):
    # 在下面的代码行中使用断点来调试脚本。
    print(f'Hi, {name}')  # 按 Ctrl+F8 切换断点。

class Auth():

    def __init__(self, app_id, app_key):
        self.app_id = app_id
        self.app_key = app_key

    def get_auth_header(self):
        content_type = 'application/x-www-form-urlencoded'
        grant_type = 'client_credentials'

        return{
            'content-type' : content_type,
            'grant_type' : grant_type,
            'client_id' : self.app_id,
            'client_secret' : self.app_key
        }

class data():

    def __init__(self, app_id, app_key, auth_response):
        self.app_id = app_id
        self.app_key = app_key
        self.auth_response = auth_response

    def get_data_header(self):
        auth_JSON = json.loads(self.auth_response.text)
        access_token = auth_JSON.get('access_token')

        return{
            'authorization': 'Bearer '+access_token
        }


# 按间距中的绿色按钮以运行脚本。
if __name__ == '__main__':
    params = {
        '$format': 'JSON',
    }
    dir ='測試.xlsx'
    wb = load_workbook(dir)
        # 獲取所有表名
    sheet_names = wb.sheetnames  # 得到工作簿的所有工作表名 結果： ['Sheet1', 'Sheet2', 'Sheet3']
        # 根據表名打開sheet表
    sheet1 = wb[sheet_names[0]]  # 打開第一個 sheet 工作表

    # 獲取N列的所有資料
    list_sheet1_column_N = []  # 請求參數

    # for b, e, c, a,f in zip(sheet1["B"], sheet1["E"], sheet1["C"], sheet1["A"],sheet1["F"]):
    for e in sheet1["N"]:
         list_sheet1_column_N.append(e.value)

    row = 2
    print(f'excel_ID: {e.value}\nexcel_type: {type(e.value)}')
    # json_data = list(e.value)
    # json_data = list_sheet1_column_N[row].split(";")
    # print(f'json_ID: {json_data}\njson_type: {type(json_data)}')
    # for method, str_param in zip(API_request, list_sheet1_column_N[1:]):
    for str_param in list_sheet1_column_N[1:]:
        json_data = str_param.strip().split(";")
        print(f'json_ID: {json_data}\njson_type: {type(json_data)}')
        try:
            a = Auth(app_id, app_key)
            auth_response = requests.post(auth_url, a.get_auth_header())

            d = data(app_id, app_key, auth_response)
            res_dist = requests.post(url_len, params=params, headers=d.get_data_header(), json=json_data)
            # print(f'res_dist: {res_dist}\nres_dist: {type(res_dist)}')
            dist_list = json.loads(res_dist.text)  ##這邊順序就變了 2023/12/18
            print(f'res_dist: {dist_list}')
            route_dist = 0
            for link in dist_list:
                route_dist += link["Length"]
                print(f'LinkID:{link["LinkID"]}\tLength:" {route_dist}')
            # ls.append(res_dist)
            print(f'route_dist: {route_dist}')

        except:
            logger.error('《' + str(row) + '》項，介面地址或入參異常！！！')
        row += 1
    try:
        wb.save(dir)
        logger.info('測試資料保存成功！！！')
    except:
        logger.error('保存失敗，可能Excel檔案未關閉，請關閉Excel檔案後重新測試')

# 访问 https://www.jetbrains.com/help/pycharm/ 获取 PyCharm 帮助
